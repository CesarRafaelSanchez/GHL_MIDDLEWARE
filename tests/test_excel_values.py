import os
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from openpyxl import load_workbook
import os
import sys

path = '/app/storage/exports/Ficha de Datos - DELLCORPS.xlsx'
if not os.path.exists(path):
    print(f"ERROR: File {path} does not exist")
    sys.exit(1)

wb = load_workbook(path, data_only=True)
ws = wb['Ficha']

print("=== VERIFYING GENERATED EXCEL VALUES ===")
print("C5 (Project Name):", ws['C5'].value)
print("C53 (Total Towers):", ws['C53'].value)
print("C54 (Total Hogares):", ws['C54'].value)

# Verify Tower A
print("\n--- TOWER A (Row 58-59) ---")
print("A58 (Tower Name Header):", ws['A58'].value)
print("B58 (Piso label):", ws['B58'].value)
print("B59 (Hogares label):", ws['B59'].value)
print("C58 (Floor 1):", ws['C58'].value)
print("C59 (Hogares Floor 1):", ws['C59'].value)
print("D58 (Floor 2):", ws['D58'].value)
print("D59 (Hogares Floor 2):", ws['D59'].value)
print("M59 (Total Formula/Value):", ws['M59'].value)

# Verify Tower B
print("\n--- TOWER B (Row 61-62) ---")
print("A61 (Tower Name Header):", ws['A61'].value)
print("C61 (Floor 1):", ws['C61'].value)
print("C62 (Hogares Floor 1):", ws['C62'].value)
print("M62 (Total Formula/Value):", ws['M62'].value)

# Verify Images
print("\n--- IMAGES ---")
print("Images count:", len(ws._images))
for i, img in enumerate(ws._images):
    print(f"Image {i+1}: cell={img.anchor}, width={img.width}, height={img.height}")

print("=== VERIFICATION COMPLETED ===")
