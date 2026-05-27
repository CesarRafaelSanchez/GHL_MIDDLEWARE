from openpyxl import load_workbook

RUTA_EXCEL = "plantillas/Ficha de Datos NOMBRE DEL PROYECTO.xlsx"

wb = load_workbook(RUTA_EXCEL)

print("📄 HOJAS ENCONTRADAS:")
for sheet_name in wb.sheetnames:
    print(f" - {sheet_name}")

print("\n🔍 CONTENIDO DE CELDAS NO VACÍAS:")
for ws in wb.worksheets:
    print(f"\n==============================")
    print(f"HOJA: {ws.title}")
    print(f"==============================")

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                print(f"{cell.coordinate}: {cell.value}")

    print("\n🔗 CELDAS COMBINADAS:")
    for merged_range in ws.merged_cells.ranges:
        print(merged_range)