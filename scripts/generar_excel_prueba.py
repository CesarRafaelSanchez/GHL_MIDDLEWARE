import os
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from openpyxl import load_workbook
from copy import copy
import os

RUTA_PLANTILLA = "resources/templates/Ficha de Datos NOMBRE DEL PROYECTO.xlsx"
CARPETA_SALIDA = "storage/exports"

os.makedirs(CARPETA_SALIDA, exist_ok=True)


def escribir(ws, celda, valor):
    """
    Escribe en una celda conservando formato.
    Si el valor viene vacío, no escribe nada.
    """
    if valor is None:
        return

    valor = str(valor).strip()
    if valor == "":
        return

    ws[celda] = valor.upper()


def marcar(ws, celda, condicion):
    """
    Marca con X una opción si la condición es verdadera.
    """
    if condicion:
        ws[celda] = "X"


def generar_excel_ficha_datos(datos):
    wb = load_workbook(RUTA_PLANTILLA)
    ws = wb["Ficha"]

    # =========================
    # DATOS GENERALES
    # =========================
    escribir(ws, "C5", datos.get("nombre_proyecto"))

    # Tipo de proyecto
    tipo_proyecto = datos.get("tipo_proyecto", "").upper()
    marcar(ws, "E8", tipo_proyecto == "NUEVO PREDIO")
    marcar(ws, "I8", tipo_proyecto == "AMPLIACION DE TORRE")

    # Fuente / origen
    fuente = datos.get("fuente_origen", "").upper()
    marcar(ws, "E12", fuente == "PROPIO")

    # Clasificación
    clasificacion = datos.get("clasificacion", "").upper()
    marcar(ws, "E16", clasificacion == "CONDOMINIO")
    marcar(ws, "I16", clasificacion == "EDIFICIO")

    # Tipo construcción
    tipo_construccion = datos.get("tipo_construccion", "").upper()
    marcar(ws, "E22", tipo_construccion == "ESTRENO")
    marcar(ws, "I22", tipo_construccion == "MODERNO")
    marcar(ws, "L22", tipo_construccion == "ANTIGUO")

    # Fechas
   # Fechas
    escribir(ws, "E23", datos.get("fecha_entrega_edificio"))
    escribir(ws, "E25", datos.get("fecha_termino_montantes"))
    escribir(ws, "E26", datos.get("fecha_termino_mecha"))

    # Junta directiva
    junta = datos.get("junta_directiva", "").upper()
    marcar(ws, "E31", junta == "SI")
    marcar(ws, "I31", junta == "NO")

    # Responsable
    escribir(ws, "D34", datos.get("cargo_responsable"))
    escribir(ws, "I34", datos.get("nombre_responsable"))
    escribir(ws, "D35", datos.get("telefono_responsable"))
    escribir(ws, "I35", datos.get("correo_responsable"))

    # Operador actual
    operador = datos.get("operador_actual", "").upper()
    marcar(ws, "E39", operador == "MOVISTAR")
    marcar(ws, "I39", operador == "NUBYX")
    marcar(ws, "L39", operador == "ENTEL")
    marcar(ws, "E40", operador == "CLARO")
    marcar(ws, "I40", operador == "WOW")
    marcar(ws, "L40", operador == "BITEL")
    marcar(ws, "E42", operador == "NINGUNO")

    # Visita técnica
    escribir(ws, "D46", datos.get("visita_inspeccion_tecnica"))

    horario = datos.get("rango_horario_visita", "").upper()
    marcar(ws, "I46", horario in ["9 AM A 12 AM", "9AM A 12M", "9 AM A 12M"])
    marcar(ws, "L46", horario == "1 PM A 4 PM")

    # Dirección
    escribir(ws, "C50", datos.get("departamento"))
    escribir(ws, "I50", datos.get("provincia"))
    escribir(ws, "C51", datos.get("distrito"))
    escribir(ws, "I51", datos.get("urbanizacion"))
    escribir(ws, "C52", datos.get("codigo_postal"))
    escribir(ws, "C53", datos.get("tipo_via"))
    escribir(ws, "C54", datos.get("nombre_via"))
    escribir(ws, "C55", datos.get("numero_via"))
    escribir(ws, "C56", datos.get("coordenadas"))

    # Datos técnicos
    escribir(ws, "C60", datos.get("total_torres"))
    escribir(ws, "C61", datos.get("total_hogares"))

    # Torre 1
    escribir(ws, "A63", f"TORRE {datos.get('nombre_torre1', '')}")
    escribir(ws, "D63", datos.get("pisos_torre1"))
    escribir(ws, "D64", datos.get("hogares_torre1"))

    # Torre 2
    escribir(ws, "A66", f"TORRE {datos.get('nombre_torre2', '')}")
    escribir(ws, "D66", datos.get("pisos_torre2"))
    escribir(ws, "D67", datos.get("hogares_torre2"))

    # Torre 3
    escribir(ws, "A69", f"TORRE {datos.get('nombre_torre3', '')}")
    escribir(ws, "D69", datos.get("pisos_torre3"))
    escribir(ws, "D70", datos.get("hogares_torre3"))

    # Clientes interesados
    escribir(ws, "C77", datos.get("clientes_interesados"))

    # Canal
    escribir(ws, "C81", datos.get("nombre_canal"))

    # Gestor
    escribir(ws, "C84", datos.get("gestor"))
    escribir(ws, "I84", datos.get("celular_gestor"))

    nombre_archivo = f"Ficha de Datos - {datos.get('nombre_proyecto', 'PROYECTO')}.xlsx"
    ruta_salida = os.path.join(CARPETA_SALIDA, nombre_archivo)

    wb.save(ruta_salida)
    return ruta_salida


if __name__ == "__main__":
    datos_prueba = {
        "nombre_proyecto": "EDIFICIO PRUEBA 31",
        "tipo_proyecto": "Nuevo Predio",
        "fuente_origen": "Propio",
        "clasificacion": "Edificio",
        "tipo_construccion": "Moderno",
        "fecha_entrega_edificio": "20/05/2026",
        "fecha_termino_montantes": "20/05/2026",
        "fecha_termino_mecha": "20/05/2026",
        "junta_directiva": "No",
        "cargo_responsable": "Ingeniero",
        "nombre_responsable": "Pedro Perez",
        "telefono_responsable": "+5193206804",
        "correo_responsable": "ejemplo@gmail.com",
        "operador_actual": "Claro",
        "visita_inspeccion_tecnica": "20/05/2026",
        "rango_horario_visita": "1 PM A 4 PM",
        "departamento": "Lima",
        "provincia": "Lima",
        "distrito": "Surco",
        "urbanizacion": "Las Casuarinas",
        "codigo_postal": "1861",
        "tipo_via": "Avenida",
        "nombre_via": "PRUEBA",
        "numero_via": "",
        "coordenadas": "123123",
        "total_torres": "79",
        "total_hogares": "8",
        "nombre_torre1": "1",
        "pisos_torre1": "8",
        "hogares_torre1": "8",
        "nombre_torre2": "2",
        "pisos_torre2": "8",
        "hogares_torre2": "8",
        "nombre_torre3": "3",
        "pisos_torre3": "8",
        "hogares_torre3": "8",
        "clientes_interesados": "8",
        "nombre_canal": "FUTURA",
        "gestor": "ALEJANDRO ABRISQUETA",
        "celular_gestor": "",
    }

    archivo = generar_excel_ficha_datos(datos_prueba)
    print(f"OK: Excel generado correctamente: {archivo}")