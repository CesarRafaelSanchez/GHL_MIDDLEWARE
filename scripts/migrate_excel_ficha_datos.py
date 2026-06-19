import os
import sys
import re
import json
import argparse
import sqlite3
import openpyxl
import requests
from datetime import datetime

# Agregar la raíz del proyecto al sys.path para poder importar app
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

def load_env_manually():
    env_vars = {}
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    env_path = os.path.join(base_dir, '.env')
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    env_vars[k.strip()] = v.strip()
    return env_vars

env = load_env_manually()
GHL_TOKEN = env.get("GHL_ACCESS_TOKEN", "pit-91efad56-ae58-41d7-85f6-6b66504b7e16")
LOCATION_ID = env.get("GHL_LOCATION_ID", "dHdydlGzW0HODg6XDQe7")
PIPELINE_ID = env.get("PIPELINE_ID", "LPfxv2sqrrNOTtjbWZ3h")

os.environ["GHL_ACCESS_TOKEN"] = GHL_TOKEN
os.environ["GHL_LOCATION_ID"] = LOCATION_ID
os.environ["PIPELINE_ID"] = PIPELINE_ID

from app.database import get_db_connection, guardar_oportunidad_db_internal
from app.services.ficha_service import construir_datos_ficha_desde_webhook
from app.utils.helpers import obtener_mapa_keys_a_ids, extraer_custom_fields_para_ghl, obtener_session_con_retries

# Mapeo de columnas del Excel a claves del middleware
EXCEL_COL_TO_MIDDLEWARE_KEY = {
    "Nombre de Canal": "cf_nombre_cancal_hunting",  # GHL fieldKey tiene typo: "cancal" no "canal"
    "Ingreso": "cf_tipo_ingreso",
    "Nombre del gestor - Teléfono": "cf_ejecutivo_principal",
    "Nombre del vendedor": "cf_gestor_real",
    "Nombre del Edificio / Condominio": "cf_nombre_proyecto",
    "Tipo de Proyecto": "cf_tipo_proyecto",
    "Fuente / Origen": "cf_fuente_hunting",
    "Clasificación": "cf_clasificacion_proyecto",
    "Tipo de construcción": "cf_tipo_construccion_edificio",
    "Fecha de entrega del Edificio (Estreno)": "cf_fecha_entrega_edificio_estreno",
    "Fecha de Termino de Montantes (Estreno)": "cf_fecha_termino_montantes_edificio_estreno",
    "Fecha de Termino de Mecha (Estreno)": "cf_fecha_termino_mecha_edificio_estreno",
    "Junta Directiva": "cf_junta_directiva",
    "Cargo del Responsable": "cf_cargo_responsable_edificio",
    "Nombre del Responsable": "cf_nombre_responsable_edificio",
    "Teléfono - Móvil": "cf_telefono_responsable_edificio",
    "Correo": "cf_correo_responsable_edificio",
    "Visita de Inspección Técnica (7 días después de la apertura)": "cf_visita_inspeccion_tecnica_win",
    "Rango Horario": "cf_rango_horario_visita_tecnica",
    "Departamento": "cf_departamento_edificio",
    "Provincia": "cf_provincia_edificio",
    "Distrito": "cf_distrito",
    "Urbanización": "cf_urbanizacion_edificio",
    "Código postal": "cf_codigo_postal_edificio",
    "Tipo de vía": "cf_tipo_via",
    "Nombre de vía": "cf_nombre_via",
    "Coordenadas (-12.1, -77.1)": "cf_coordenadas",
    "Total de Torres": "cf_total_torres_proyecto",
    "Total Hogares": "cf_total_hogares_proyecto",
    "Nombre Primera Torre (1; A; Nombre)": "cf_nombre_torre1_proyecto",
    "Cantidad de Pisos (Primera Torre)": "cf_pisos_torre1_proyecto",
    "Cantidad de Hogares Por Piso (Primera Torre)": "cf_hogares_por_piso_torre1",
    "Nombre Segunda Torre (2; B; Nombre)": "cf_nombre_torre2_proyecto",
    "Cantidad de Pisos (Segunda Torre)": "cf_pisos_torre2_proyecto",
    "Cantidad de Hogares por Piso(Segunda Torre)": "cf_hogares_por_piso_torre2",
    "Nombre Tercera Torre (3; C; Nombre)": "cf_nombre_torre3_proyecto",
    "Cantidad de Pisos (Tercera Torre)": "cf_pisos_torre3_proyecto",
    "Cantidad de Hogares por Piso (Tercera Torre)": "cf_hogares_por_piso_torre3",
    "Clientes interesados": "cf_cantidad_clientes_interesados",
    "Foto del edificio": "cf_foto_edificio",
    "Foto de las montantes (ducterías) y acometida (mecha)  ": "cf_foto_montantes",
    "Numeración de vía": "cf_numeracion_via"
}

STAGE_HABILITACION_COMPLETA = "dc5a218f-50a8-4bb6-9351-82b2f10d9886"

def normalize_text(text):
    if not text:
        return ""
    text = str(text).strip().upper()
    # Remover prefijos comunes
    if text.startswith("ADMINISTRACION:") or text.startswith("ADMINISTRACIÓN:"):
        text = text.split(":", 1)[1].strip()
    # Reemplazar acentos
    replacements = {"Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U", "Ü": "U", "Ñ": "N"}
    for k, v in replacements.items():
        text = text.replace(k, v)
    # Remover prefijos de tipo de predio (Edificio, Condominio, etc.)
    text = re.sub(r"^(EDIFICIO|CONDOMINIO|EDF|RESIDENCIAL)\s+", "", text)
    # Remover tabulaciones y espacios dobles
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def extract_gdrive_id(url):
    if not url or not isinstance(url, str):
        return None
    match1 = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
    if match1:
        return match1.group(1)
    match2 = re.search(r"/file/d/([a-zA-Z0-9_-]+)", url)
    if match2:
        return match2.group(1)
    return None

def get_gdrive_download_url(file_id):
    return f"https://drive.google.com/uc?export=download&id={file_id}"

def download_image(url, output_path, dry_run=False):
    file_id = extract_gdrive_id(url)
    download_url = get_gdrive_download_url(file_id) if file_id else url
    if dry_run:
        print(f"[DRY-RUN] Descargaría imagen desde {download_url} hacia {output_path}")
        return True
    try:
        session = obtener_session_con_retries()
        resp = session.get(download_url, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code == 200:
            ct = resp.headers.get("Content-Type", "").lower()
            if "html" in ct or "text" in ct:
                print(f"⚠️  El archivo de Drive requiere inicio de sesión (no es público): {url}")
                return False
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, "wb") as f:
                f.write(resp.content)
            return True
        else:
            print(f"⚠️  Error HTTP {resp.status_code} al descargar: {url}")
            return False
    except Exception as e:
        print(f"⚠️  Excepción al descargar {url}: {e}")
        return False

def clean_filename(name):
    name = normalize_text(name).lower()
    name = re.sub(r"[^a-z0-9_-]", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name

def main():
    parser = argparse.ArgumentParser(description="Migrar Fichas de Datos desde Excel a GHL y SQLite")
    parser.add_argument("--excel", default="CONSOLIDADO DE FICHAS - WIN (respuestas) (1).xlsx", help="Ruta del Excel")
    parser.add_argument("--dry-run", action="store_true", help="Simular la migración sin escribir cambios")
    parser.add_argument("--limit", type=int, default=None, help="Límite de filas a procesar (para testing)")
    args = parser.parse_args()

    print(f"🚀 Iniciando migración de Excel Fichas de Datos...")
    print(f"📌 Archivo Excel: {args.excel}")
    print(f"📌 Modo Simulación (Dry-run): {args.dry_run}")
    if args.limit:
        print(f"📌 Límite de procesamiento: {args.limit} filas")

    # 1. Cargar base de datos local para emparejamiento rápido
    print("[INFO] Cargando base de datos SQLite del Middleware...")
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("SELECT oportunidad_id, contacto_id, nombre_proyecto, gestor, raw_json FROM oportunidades_ficha")
    db_rows = cursor.fetchall()
    
    # Mapear por nombre normalizado
    # Para manejar duplicados, agrupamos por nombre normalizado
    db_by_name = {}
    for r in db_rows:
        norm = normalize_text(r["nombre_proyecto"])
        if norm not in db_by_name:
            db_by_name[norm] = []
        db_by_name[norm].append(r)
    
    print(f"[INFO] Se cargaron {len(db_rows)} oportunidades de la caché SQLite.")

    # 2. Cargar estado de habilitados desde la hoja 'REPORTE DE FICHAS ENVIADAS'
    print("[INFO] Cargando estados de habilitación desde 'REPORTE DE FICHAS ENVIADAS'...")
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    
    reporte_sheet = wb['REPORTE DE FICHAS ENVIADAS']
    reporte_headers = [c.value for c in reporte_sheet[1]]
    idx_predio = reporte_headers.index('NOMBRE DE PREDIO')
    idx_substatus = reporte_headers.index('SUB STATUS')
    
    habilitados_set = set()
    for row in reporte_sheet.iter_rows(min_row=2, values_only=True):
        if not row[idx_predio]:
            continue
        predio_norm = normalize_text(row[idx_predio])
        substatus = str(row[idx_substatus]).strip().upper() if row[idx_substatus] else ""
        if substatus == "EDIFICIO HABILITADO":
            habilitados_set.add(predio_norm)
            
    print(f"[INFO] Se encontraron {len(habilitados_set)} edificios marcados como 'EDIFICIO HABILITADO' en el reporte.")

    # 3. Procesar hoja de respuestas
    respuestas_sheet = wb['Respuestas de formulario 1']
    respuestas_headers = [c.value for c in respuestas_sheet[1]]
    
    # Construir mapa de índices de columnas
    col_to_idx = {}
    for col_name, middleware_key in EXCEL_COL_TO_MIDDLEWARE_KEY.items():
        if col_name in respuestas_headers:
            col_to_idx[middleware_key] = respuestas_headers.index(col_name)
        else:
            # Buscar coincidencia parcial si hay espacios o caracteres especiales
            matched_idx = -1
            for idx, h in enumerate(respuestas_headers):
                if h and col_name.strip().lower() == h.strip().lower():
                    matched_idx = idx
                    break
            if matched_idx != -1:
                col_to_idx[middleware_key] = matched_idx
            else:
                print(f"⚠️  Columna del Excel '{col_name}' no encontrada en las cabeceras.")

    print(f"[INFO] Cabeceras mapeadas exitosamente. Iniciando procesamiento de filas...")

    session = obtener_session_con_retries()
    headers_contact = {
        "Authorization": f"Bearer {GHL_TOKEN}",
        "Version": "2021-04-15",
        "Content-Type": "application/json"
    }
    headers_opp = {
        "Authorization": f"Bearer {GHL_TOKEN}",
        "Version": "2021-04-15",
        "Content-Type": "application/json"
    }

    matched_count = 0
    skipped_count = 0
    stage_moved_count = 0
    download_success_count = 0
    download_failed_count = 0

    processed_rows = 0
    for r_idx, row in enumerate(respuestas_sheet.iter_rows(min_row=2, values_only=True), 2):
        if args.limit and processed_rows >= args.limit:
            break
        
        # Obtener nombre del edificio
        idx_name = col_to_idx.get("cf_nombre_proyecto")
        if idx_name is None or not row[idx_name]:
            continue
        
        building_name_raw = row[idx_name]
        building_name_norm = normalize_text(building_name_raw)
        
        # Intentar emparejar
        matches = db_by_name.get(building_name_norm, [])
        matched_db_row = None
        
        if len(matches) == 1:
            matched_db_row = matches[0]
        elif len(matches) > 1:
            # Disambiguación por gestor
            idx_vendedor = col_to_idx.get("cf_gestor_real")
            excel_gestor = normalize_text(row[idx_vendedor]) if idx_vendedor is not None and row[idx_vendedor] else ""
            
            for m in matches:
                db_gestor = normalize_text(m["gestor"])
                if excel_gestor and (excel_gestor in db_gestor or db_gestor in excel_gestor):
                    matched_db_row = m
                    break
            if not matched_db_row:
                matched_db_row = matches[0]  # Fallback al primero
                
        if not matched_db_row:
            print(f"❌ Fila {r_idx} | No se encontró coincidencia en BD para: '{building_name_raw}'")
            skipped_count += 1
            continue
            
        processed_rows += 1
        opp_id = matched_db_row["oportunidad_id"]
        contact_id = matched_db_row["contacto_id"]
        matched_count += 1
        
        print(f"✅ Fila {r_idx} | Emparejado '{building_name_raw}' -> Opp: {opp_id} | Contact: {contact_id}")

        # Extraer campos de la fila del Excel y construir flat_fields
        flat_fields = {}
        for key, idx in col_to_idx.items():
            val = row[idx]
            if val is not None and str(val).strip() not in ["", "-"]:
                if hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                flat_fields[key] = val

        # Procesar fotos
        foto_edificio_url = flat_fields.get("cf_foto_edificio")
        foto_montantes_url = flat_fields.get("cf_foto_montantes")
        
        clean_proj_name = clean_filename(building_name_raw)
        
        foto_edificio_path = None
        foto_montantes_path = None
        
        if foto_edificio_url and str(foto_edificio_url).startswith("http"):
            local_path = os.path.join("storage", "temp", f"{clean_proj_name}_foto_edificio.jpg")
            abs_local_path = os.path.abspath(local_path)
            if download_image(foto_edificio_url, abs_local_path, args.dry_run):
                foto_edificio_path = local_path
                download_success_count += 1
            else:
                download_failed_count += 1
                
        if foto_montantes_url and str(foto_montantes_url).startswith("http"):
            local_path = os.path.join("storage", "temp", f"{clean_proj_name}_foto_montantes.jpg")
            abs_local_path = os.path.abspath(local_path)
            if download_image(foto_montantes_url, abs_local_path, args.dry_run):
                foto_montantes_path = local_path
                download_success_count += 1
            else:
                download_failed_count += 1

        # Construir mapeo estructurado para actualizar base de datos local
        datos_mapeados = construir_datos_ficha_desde_webhook(flat_fields)
        if foto_edificio_path:
            datos_mapeados["foto_edificio_path"] = foto_edificio_path
        if foto_montantes_path:
            datos_mapeados["foto_montantes_path"] = foto_montantes_path

        # 4. Actualizar GoHighLevel
        custom_fields_ghl = extraer_custom_fields_para_ghl(flat_fields)
        
        if not args.dry_run:
            # A. Actualizar contacto en GHL (Campos del Formulario 2)
            # Primero mapeamos responsable nombre a firstName/lastName
            resp_nombre = flat_fields.get("cf_nombre_responsable_edificio")
            resp_tel = flat_fields.get("cf_telefono_responsable_edificio")
            resp_correo = flat_fields.get("cf_correo_responsable_edificio")
            
            contact_payload = {
                "customFields": custom_fields_ghl
            }
            if resp_nombre:
                parts = str(resp_nombre).strip().split(" ", 1)
                contact_payload["firstName"] = parts[0]
                contact_payload["lastName"] = parts[1] if len(parts) > 1 else ""
            if resp_tel:
                contact_payload["phone"] = resp_tel
            if resp_correo and "@" in str(resp_correo) and "." in str(resp_correo):
                contact_payload["email"] = str(resp_correo).strip()

            res_c = session.put(f"https://services.leadconnectorhq.com/contacts/{contact_id}", headers=headers_contact, json=contact_payload)
            if res_c.status_code == 200:
                print(f"   ↳ [GHL] Contacto {contact_id} actualizado.")
            else:
                print(f"   ↳ ⚠️  [GHL ERROR] Contacto {contact_id} no actualizado: {res_c.text}")

        # B. Comprobar si se cambia a Habilitación Completa
        mover_etapa = building_name_norm in habilitados_set
        if mover_etapa:
            print(f"   ↳ [STAGE] Edificio marcado como HABILITADO. Cambiando etapa en GHL a Habilitación Completa...")
            stage_moved_count += 1
            if not args.dry_run:
                res_o = session.put(
                    f"https://services.leadconnectorhq.com/opportunities/{opp_id}", 
                    headers=headers_opp,
                    json={"pipelineId": PIPELINE_ID, "pipelineStageId": STAGE_HABILITACION_COMPLETA}
                )
                if res_o.status_code == 200:
                    print(f"   ↳ [GHL] Oportunidad {opp_id} movida a etapa Habilitación Completa.")
                else:
                    print(f"   ↳ ⚠️  [GHL ERROR] Oportunidad {opp_id} no movida de etapa: {res_o.text}")

        # 5. Sincronizar Base de Datos Local
        if not args.dry_run:
            # Obtener el raw_json original y actualizarlo para mantener el caché íntegro
            raw_json_dict = {}
            if matched_db_row["raw_json"]:
                try:
                    raw_json_dict = json.loads(matched_db_row["raw_json"])
                except Exception:
                    pass
            # Actualizar raw_json_dict con los campos planos del Excel
            raw_json_dict.update(flat_fields)
            
            # Insertar / Actualizar
            guardar_oportunidad_db_internal(cursor, opp_id, contact_id, datos_mapeados, raw_json_dict)
            print(f"   ↳ [SQLITE] Datos actualizados en caché local para opp {opp_id}")

    if not args.dry_run:
        conn.commit()
    conn.close()

    print("\n" + "="*50)
    print("📋 RESUMEN DE EJECUCIÓN:")
    print(f"  • Filas procesadas: {processed_rows}")
    print(f"  • Coincidencias emparejadas con éxito: {matched_count}")
    print(f"  • Filas omitidas (no encontradas en BD): {skipped_count}")
    print(f"  • Oportunidades promovidas a Habilitación Completa: {stage_moved_count}")
    print(f"  • Imágenes descargadas exitosamente: {download_success_count}")
    print(f"  • Descargas de imágenes fallidas (privados/errores): {download_failed_count}")
    print("="*50 + "\n")

if __name__ == "__main__":
    main()
