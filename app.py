import os
import requests
from flask import Flask, request, jsonify
from dotenv import load_dotenv
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import smtplib
from email.message import EmailMessage
from urllib.parse import urlparse
from io import BytesIO
load_dotenv()

app = Flask(__name__)


# Credenciales y Configuración
GHL_TOKEN = os.getenv("GHL_ACCESS_TOKEN")
LOCATION_ID = os.getenv("GHL_LOCATION_ID")
PIPELINE_ID = os.getenv("PIPELINE_ID")
STAGE_FORM_1 = os.getenv("STAGE_FORM_1_COMPLETADO")
STAGE_FORM_2 = os.getenv("STAGE_FORM_2_COMPLETADO")
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = os.getenv("SMTP_PORT")
EMAIL_EMISOR = os.getenv("EMAIL_EMISOR")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
EMAIL_DESTINO = os.getenv("EMAIL_DESTINO")
EMAIL_CC_FICHA = os.getenv("EMAIL_CC_FICHA")
EMAIL_CC_ASIGNACION = os.getenv("EMAIL_CC_ASIGNACION")
STAGE_PENDIENTE_INICIO_HABILITACION = os.getenv("STAGE_PENDIENTE_INICIO_HABILITACION")

HEADERS_GHL = {
    "Authorization": f"Bearer {GHL_TOKEN}",
    "Version": "2021-04-15",
    "Content-Type": "application/json"
}

USERS_GHL = {
    "JEAN": "UzEVMjDvEHlw6YUAj3aJ",
    "YASMIN": "bVGkAziqy6vwDoFbqvr6",
    "STEFANO_BO": "6u8iZhDXnxp0xSp7XfDl"
}

RUTA_PLANTILLA_FICHA_DATOS = "plantillas/Ficha de Datos NOMBRE DEL PROYECTO.xlsx"
CARPETA_SALIDAS = "salidas"

os.makedirs(CARPETA_SALIDAS, exist_ok=True)
CARPETA_TEMP = "temp"
os.makedirs(CARPETA_TEMP, exist_ok=True)

ARCHIVO_CACHE_FICHAS = "cache_fichas_datos.json"


def normalizar_clave_proyecto(nombre):
    return str(nombre or "").strip().upper()


def cargar_cache_fichas():
    if not os.path.exists(ARCHIVO_CACHE_FICHAS):
        return {}

    try:
        with open(ARCHIVO_CACHE_FICHAS, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}


def guardar_cache_fichas(cache):
    with open(ARCHIVO_CACHE_FICHAS, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def guardar_datos_ficha_en_cache(nombre_proyecto, opp_id, contact_id, datos):
    cache = cargar_cache_fichas()

    clave_nombre = normalizar_clave_proyecto(nombre_proyecto)

    datos_guardar = dict(datos)
    datos_guardar["_opp_id"] = opp_id
    datos_guardar["_contact_id"] = contact_id
    datos_guardar["_nombre_proyecto"] = clave_nombre
    datos_guardar["_guardado_en"] = datetime.now().isoformat()

    if clave_nombre:
        cache[f"nombre::{clave_nombre}"] = datos_guardar

    if opp_id:
        cache[f"opp::{opp_id}"] = datos_guardar

    if contact_id:
        cache[f"contact::{contact_id}"] = datos_guardar

    guardar_cache_fichas(cache)

    print(f"💾 [CACHE FICHA] Datos guardados para: {clave_nombre}")


def recuperar_datos_ficha_de_cache(data):
    cache = cargar_cache_fichas()

    opp_id = data.get("id")
    contact_id = data.get("contact_id") or data.get("contact", {}).get("id")
    nombre = data.get("cf_nombre_proyecto") or data.get("opportunity_name")
    clave_nombre = normalizar_clave_proyecto(nombre)

    posibles_claves = []

    if opp_id:
        posibles_claves.append(f"opp::{opp_id}")

    if contact_id:
        posibles_claves.append(f"contact::{contact_id}")

    if clave_nombre:
        posibles_claves.append(f"nombre::{clave_nombre}")

    for clave in posibles_claves:
        if clave in cache:
            print(f"✅ [CACHE FICHA] Datos recuperados usando clave: {clave}")

            datos_cache = cache[clave]

            # Mezclamos: lo del cache manda, pero conservamos ids actuales del webhook final
            data_mezclada = dict(data)
            data_mezclada.update(datos_cache)

            data_mezclada["id"] = opp_id or datos_cache.get("_opp_id")
            data_mezclada["contact_id"] = contact_id or datos_cache.get("_contact_id")

            return data_mezclada

    print("⚠️ [CACHE FICHA] No se encontraron datos completos en cache.")
    return data
def escribir_excel(ws, celda, valor):
    if valor is None:
        return

    valor = str(valor).strip()
    if valor == "":
        return

    ws[celda] = valor.upper()


def marcar_excel(ws, celda, condicion):
    if condicion:
        ws[celda] = "X"
def obtener_primera_url(valor):
    """
    GHL puede mandar la foto como lista o como texto.
    Esta función devuelve solo la primera URL.
    """
    if not valor:
        return ""

    if isinstance(valor, list) and len(valor) > 0:
        return valor[0]

    if isinstance(valor, str):
        return valor

    return ""
def limpiar_nombre_archivo(nombre):
    nombre = str(nombre or "archivo")
    caracteres_invalidos = '<>:"/\\|?*'
    for c in caracteres_invalidos:
        nombre = nombre.replace(c, "_")
    return nombre.strip().replace(" ", "_")

def descargar_imagen(url, nombre_archivo):
    """
    Descarga imagen desde GHL documents/download usando PIT token.
    Convierte cualquier imagen a JPG estándar para evitar errores tipo .mpo en Excel.
    """
    if not url:
        print("⚠️ No llegó URL de imagen.")
        return None

    try:
        nombre_archivo = limpiar_nombre_archivo(nombre_archivo)

        headers = {
            "Authorization": f"Bearer {GHL_TOKEN}",
            "Version": "2021-04-15",
            "Accept": "*/*"
        }

        print(f"📸 Intentando descargar imagen GHL: {url}")

        response = requests.get(
            url,
            headers=headers,
            timeout=30,
            allow_redirects=True
        )

        print("📸 Status descarga imagen:", response.status_code)
        print("📸 Content-Type:", response.headers.get("Content-Type"))

        if response.status_code != 200:
            print("❌ Error descargando imagen:")
            print(response.text[:500])
            return None

        # Convertimos SIEMPRE a JPG estándar
        imagen = PILImage.open(BytesIO(response.content))

        # Algunos archivos vienen como MPO/JPEG raro.
        # Convertimos a RGB para que Excel/openpyxl lo acepte.
        if imagen.mode in ("RGBA", "P"):
            imagen = imagen.convert("RGB")
        else:
            imagen = imagen.convert("RGB")

        ruta = os.path.join(CARPETA_TEMP, f"{nombre_archivo}.jpg")

        imagen.save(ruta, format="JPEG", quality=90)

        print(f"✅ Imagen descargada y convertida correctamente: {ruta}")
        return ruta

    except Exception as e:
        print(f"❌ Error descargando/convirtiendo imagen GHL: {e}")
        return None

def insertar_imagen_excel(ws, ruta_imagen, celda, ancho=330, alto=260):
    if not ruta_imagen:
        print(f"⚠️ No hay imagen para insertar en {celda}")
        return

    if not os.path.exists(ruta_imagen):
        print(f"⚠️ No existe la imagen local: {ruta_imagen}")
        return

    try:
        img = ExcelImage(ruta_imagen)
        img.width = ancho
        img.height = alto
        ws.add_image(img, celda)
        print(f"✅ Imagen insertada en Excel: {ruta_imagen} -> {celda}")
    except Exception as e:
        print(f"❌ Error insertando imagen en Excel: {e}")      


def generar_excel_ficha_datos(datos):
    wb = load_workbook(RUTA_PLANTILLA_FICHA_DATOS)
    ws = wb["Ficha"]

    # DATOS GENERALES
    escribir_excel(ws, "C5", datos.get("nombre_proyecto"))

    # Tipo de proyecto
    tipo_proyecto = datos.get("tipo_proyecto", "").upper()
    marcar_excel(ws, "E8", tipo_proyecto == "NUEVO PREDIO")
    marcar_excel(ws, "I8", tipo_proyecto == "AMPLIACION DE TORRE")

    # Fuente / origen
    fuente = datos.get("fuente_origen", "").upper()
    marcar_excel(ws, "E12", fuente == "PROPIO")

    # Clasificación
    clasificacion = datos.get("clasificacion", "").upper()
    marcar_excel(ws, "E16", clasificacion == "CONDOMINIO")
    marcar_excel(ws, "I16", clasificacion == "EDIFICIO")

    # Tipo construcción
    tipo_construccion = datos.get("tipo_construccion", "").upper()
    marcar_excel(ws, "E22", tipo_construccion == "ESTRENO")
    marcar_excel(ws, "I22", tipo_construccion == "MODERNO")
    marcar_excel(ws, "L22", tipo_construccion == "ANTIGUO")

    # Fechas
    escribir_excel(ws, "E23", datos.get("fecha_entrega_edificio"))
    escribir_excel(ws, "E25", datos.get("fecha_termino_montantes"))
    escribir_excel(ws, "E26", datos.get("fecha_termino_mecha"))

    # Junta directiva
    junta = datos.get("junta_directiva", "").upper()
    marcar_excel(ws, "E31", junta == "SI")
    marcar_excel(ws, "I31", junta == "NO")

    # Responsable
    escribir_excel(ws, "D34", datos.get("cargo_responsable"))
    escribir_excel(ws, "I34", datos.get("nombre_responsable"))
    escribir_excel(ws, "D35", datos.get("telefono_responsable"))
    escribir_excel(ws, "I35", datos.get("correo_responsable"))

    # Operador actual
    operador = datos.get("operador_actual", "").upper()
    marcar_excel(ws, "E39", operador == "MOVISTAR")
    marcar_excel(ws, "I39", operador == "NUBYX")
    marcar_excel(ws, "L39", operador == "ENTEL")
    marcar_excel(ws, "E40", operador == "CLARO")
    marcar_excel(ws, "I40", operador == "WOW")
    marcar_excel(ws, "L40", operador == "BITEL")
    marcar_excel(ws, "E42", operador == "NINGUNO")

    # Visita técnica
    escribir_excel(ws, "D46", datos.get("visita_inspeccion_tecnica"))

    horario = datos.get("rango_horario_visita", "").upper()
    marcar_excel(ws, "I46", horario in ["9 AM A 12 AM", "9AM A 12M", "9 AM A 12M"])
    marcar_excel(ws, "L46", horario == "1 PM A 4 PM")

    # Dirección
    escribir_excel(ws, "C50", datos.get("departamento"))
    escribir_excel(ws, "I50", datos.get("provincia"))
    escribir_excel(ws, "C51", datos.get("distrito"))
    escribir_excel(ws, "I51", datos.get("urbanizacion"))
    escribir_excel(ws, "C52", datos.get("codigo_postal"))
    escribir_excel(ws, "C53", datos.get("tipo_via"))
    escribir_excel(ws, "C54", datos.get("nombre_via"))
    escribir_excel(ws, "C55", datos.get("numero_via"))
    escribir_excel(ws, "C56", datos.get("coordenadas"))

    # Datos técnicos
    escribir_excel(ws, "C60", datos.get("total_torres"))
    escribir_excel(ws, "C61", datos.get("total_hogares"))

    # Torre 1
    escribir_excel(ws, "A63", f"TORRE {datos.get('nombre_torre1', '')}")
    escribir_excel(ws, "D63", datos.get("pisos_torre1"))
    escribir_excel(ws, "D64", datos.get("hogares_torre1"))

    # Torre 2
    escribir_excel(ws, "A66", f"TORRE {datos.get('nombre_torre2', '')}")
    escribir_excel(ws, "D66", datos.get("pisos_torre2"))
    escribir_excel(ws, "D67", datos.get("hogares_torre2"))

    # Torre 3
    escribir_excel(ws, "A69", f"TORRE {datos.get('nombre_torre3', '')}")
    escribir_excel(ws, "D69", datos.get("pisos_torre3"))
    escribir_excel(ws, "D70", datos.get("hogares_torre3"))

    # Clientes interesados
    escribir_excel(ws, "C77", datos.get("clientes_interesados"))

    # Canal
    escribir_excel(ws, "C81", datos.get("nombre_canal"))

    # Gestor
    escribir_excel(ws, "C84", datos.get("gestor"))
    escribir_excel(ws, "I84", datos.get("celular_gestor"))

    # Fotos - insertar imágenes reales descargadas desde GHL
    insertar_imagen_excel(
        ws,
        datos.get("foto_edificio_path"),
        "A89",
        ancho=330,
        alto=260
    )

    insertar_imagen_excel(
        ws,
        datos.get("foto_montantes_path"),
        "F89",
        ancho=330,
        alto=260
    )

    nombre_proyecto = limpiar_nombre_archivo(datos.get("nombre_proyecto", "PROYECTO")).replace("_", " ")
    nombre_archivo = f"Ficha de Datos - {nombre_proyecto}.xlsx"
    ruta_salida = os.path.join(CARPETA_SALIDAS, nombre_archivo)

    wb.save(ruta_salida)
    return ruta_salida

def enviar_correo_ficha_datos_win(archivo_excel, datos):
    proyecto = datos.get("nombre_proyecto", "PROYECTO")

    asunto = f"FICHA DE DATOS {proyecto}"

    cuerpo_html = """
    <html>
      <body style="font-family: Arial, sans-serif; color: #333; font-size: 14px; line-height: 1.5;">
        <p>Buenas tardes, hacemos envío de la ficha de datos del predio en mención.</p>

        <br>

        <p>Saludos,</p>

        <br>

        <p>
          Stefano Sotomarino Goche<br>
          Back Office - Futura
        </p>
      </body>
    </html>
    """

    cuerpo_texto = """
Buenas tardes, hacemos envío de la ficha de datos del predio en mención.

Saludos,

Stefano Sotomarino Goche
Back Office - Futura
"""

    msg = EmailMessage()
    msg["From"] = EMAIL_EMISOR
    msg["To"] = EMAIL_DESTINO

    if EMAIL_CC_FICHA:
        msg["Cc"] = EMAIL_CC_FICHA

    msg["Subject"] = asunto

    msg.set_content(cuerpo_texto)
    msg.add_alternative(cuerpo_html, subtype="html")

    with open(archivo_excel, "rb") as f:
        contenido_excel = f.read()

    nombre_adjunto = os.path.basename(archivo_excel)

    msg.add_attachment(
        contenido_excel,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=nombre_adjunto
    )

    destinatarios = [EMAIL_DESTINO]

    if EMAIL_CC_FICHA:
        destinatarios.append(EMAIL_CC_FICHA)

    with smtplib.SMTP_SSL(SMTP_SERVER, int(SMTP_PORT)) as smtp:
        smtp.login(EMAIL_EMISOR, EMAIL_PASSWORD)
        smtp.send_message(msg, to_addrs=destinatarios)

    print(f"📧 [FICHA DATOS WIN] Correo enviado con adjunto: {nombre_adjunto}")

    if EMAIL_CC_FICHA:
        print(f"📧 [FICHA DATOS WIN] CC enviado a: {EMAIL_CC_FICHA}")

from datetime import datetime


def formatear_fecha(fecha):
    """
    Convierte fechas tipo 2026-09-09 a 09/09/2026.
    Si viene vacía, retorna vacío.
    """
    if not fecha:
        return ""

    try:
        return datetime.strptime(str(fecha), "%Y-%m-%d").strftime("%d/%m/%Y")
    except:
        return str(fecha)


def construir_datos_ficha_desde_webhook(data):
    """
    Convierte los campos que manda GHL al formato que usa generar_excel_ficha_datos().
    """

    datos = {
        "nombre_proyecto": data.get("cf_nombre_proyecto") or data.get("opportunity_name"),

        "tipo_proyecto": data.get("cf_tipo_proyecto"),
        "fuente_origen": data.get("cf_fuente_hunting"),
        "clasificacion": "Condominio" if "condominio" in str(data.get("cf_clasificacion_proyecto", "")).lower() else "Edificio",

        "tipo_construccion": (
            "Estreno" if "estreno" in str(data.get("cf_tipo_construccion_edificio", "")).lower()
            else "Moderno" if "moderno" in str(data.get("cf_tipo_construccion_edificio", "")).lower()
            else "Antiguo" if "antiguo" in str(data.get("cf_tipo_construccion_edificio", "")).lower()
            else data.get("cf_tipo_construccion_edificio")
        ),

        "fecha_entrega_edificio": formatear_fecha(
            data.get("cf_fecha_entrega_edificio_estreno") or data.get("cf_fecha_entrega_edificio")
        ),
        "fecha_termino_montantes": formatear_fecha(
            data.get("cf_fecha_termino_montantes_edificio_estreno")
        ),
        "fecha_termino_mecha": formatear_fecha(
            data.get("cf_fecha_termino_mecha_edificio_estreno")
        ),

        "junta_directiva": data.get("cf_junta_directiva"),
        "cargo_responsable": data.get("cf_cargo_responsable_edificio") or data.get("cf_cargo_responsable"),
        "nombre_responsable": data.get("cf_nombre_responsable_edificio"),
        "telefono_responsable": data.get("cf_telefono_responsable_edificio"),
        "correo_responsable": data.get("cf_correo_responsable_edificio"),

        "operador_actual": data.get("cf_operador_actual"),

        "visita_inspeccion_tecnica": formatear_fecha(data.get("cf_visita_inspeccion_tecnica_win")),
        "rango_horario_visita": data.get("cf_rango_horario_visita_tecnica"),

        "departamento": data.get("cf_departamento_edificio"),
        "provincia": data.get("cf_provincia_edificio"),
        "distrito": data.get("cf_distrito"),
        "urbanizacion": data.get("cf_urbanizacion_edificio"),
        "codigo_postal": data.get("cf_codigo_postal_edificio"),
        "tipo_via": data.get("cf_tipo_via"),
        "nombre_via": data.get("cf_nombre_via"),
        "numero_via": data.get("cf_numeracion_via"),
        "coordenadas": data.get("cf_coordenadas"),

        "total_torres": data.get("cf_total_torres_proyecto"),
        "total_hogares": data.get("cf_total_hogares_proyecto"),

        "nombre_torre1": data.get("cf_nombre_torre1_proyecto"),
        "pisos_torre1": data.get("cf_pisos_torre1_proyecto"),
        "hogares_torre1": data.get("cf_hogares_torre1_proyecto"),

        "nombre_torre2": data.get("cf_nombre_torre2_proyecto"),
        "pisos_torre2": data.get("cf_pisos_torre2_proyecto"),
        "hogares_torre2": data.get("cf_hogares_torre2_proyecto"),

        "nombre_torre3": data.get("cf_nombre_torre3_proyecto"),
        "pisos_torre3": data.get("cf_pisos_torre3_proyecto"),
        "hogares_torre3": data.get("cf_hogares_torre3_proyecto"),

        "clientes_interesados": data.get("cf_cantidad_clientes_interesados"),
        "nombre_canal": data.get("cf_nombre_cancal_hunting") or data.get("cf_nombre_canal_hunting"),

        "gestor": data.get("cf_gestor_real"),
        "celular_gestor": data.get("user", {}).get("phone", ""),

        # Fotos que llegan desde GHL
        "foto_edificio": obtener_primera_url(data.get("cf_foto_edificio")),
        "foto_montantes": obtener_primera_url(data.get("cf_foto_montantes")),
    }

    print("📸 [MAPEO] cf_foto_edificio:", data.get("cf_foto_edificio"))
    print("📸 [MAPEO] cf_foto_montantes:", data.get("cf_foto_montantes"))
    print("📸 [MAPEO] foto_edificio final:", datos.get("foto_edificio"))
    print("📸 [MAPEO] foto_montantes final:", datos.get("foto_montantes"))

    return datos


def obtener_campo(payload, key):
    if key in payload: return payload[key]
    cf = payload.get("customFields", {})
    if isinstance(cf, dict) and key in cf: return cf[key]
    return None


def extraer_custom_fields_para_ghl(datos_formulario):
    cf_array = []
    fuente = datos_formulario.get("customFields", datos_formulario)
    if isinstance(fuente, dict):
        for k, v in fuente.items():
            if k.startswith("cf_") and v is not None and v != "":
                cf_array.append({"id": k, "value": v})
    return cf_array
def inyectar_custom_fields_desde_contacto(data):
    """
    Si el webhook final llega sin campos personalizados,
    consulta el contacto en GHL y mete los custom fields dentro de data.
    """
    contact_id = data.get("contact_id") or data.get("contact", {}).get("id")

    if not contact_id:
        print("⚠️ [FICHA DATOS WIN] No llegó contact_id para recuperar custom fields.")
        return data

    try:
        print(f"🔎 [FICHA DATOS WIN] Recuperando custom fields del contacto: {contact_id}")

        res_contact = requests.get(
            f"https://services.leadconnectorhq.com/contacts/{contact_id}",
            headers=HEADERS_GHL
        )

        print("🔎 [FICHA DATOS WIN] Status GET contacto:", res_contact.status_code)

        if res_contact.status_code != 200:
            print("❌ [FICHA DATOS WIN] Error GET contacto:", res_contact.text[:500])
            return data

        contacto = res_contact.json().get("contact", {})
        custom_fields = contacto.get("customFields", [])

        for cf in custom_fields:
            cf_id = cf.get("id")
            cf_value = cf.get("value")

            if cf_id and cf_value not in [None, ""]:
                data[cf_id] = cf_value

        print(f"✅ [FICHA DATOS WIN] Custom fields inyectados: {len(custom_fields)}")
        return data

    except Exception as e:
        print("❌ [FICHA DATOS WIN] Error recuperando custom fields:", str(e))
        return data

def mover_oportunidad_a_pendiente_inicio_habilitacion(opp_id):
    if not opp_id:
        print("⚠️ [FICHA DATOS WIN] No llegó opportunity ID para mover etapa.")
        return False

    if not STAGE_PENDIENTE_INICIO_HABILITACION:
        print("⚠️ [FICHA DATOS WIN] Falta STAGE_PENDIENTE_INICIO_HABILITACION en .env")
        return False

    try:
        res = requests.put(
            f"https://services.leadconnectorhq.com/opportunities/{opp_id}",
            headers=HEADERS_GHL,
            json={
                "pipelineId": PIPELINE_ID,
                "pipelineStageId": STAGE_PENDIENTE_INICIO_HABILITACION
            }
        )

        if res.status_code == 200:
            print("✅ [FICHA DATOS WIN] Tarjeta movida a 'Pendiente Inicio de Habilitación'.")
            return True

        print("❌ [FICHA DATOS WIN] Error moviendo tarjeta:", res.text[:500])
        return False

    except Exception as e:
        print("❌ [FICHA DATOS WIN] Excepción moviendo tarjeta:", str(e))
        return False

@app.route("/webhook-enviar-ficha-datos-win", methods=["POST"])
def webhook_enviar_ficha_datos_win():
    try:
        data = request.get_json(silent=True) or request.form.to_dict()

        # Primero recuperamos la ficha completa guardada desde /webhook-formulario2.
        # Esto evita que el Excel salga vacío cuando el webhook final llega incompleto.
        data = recuperar_datos_ficha_de_cache(data)

        print("\n📩 [FICHA DATOS WIN] Webhook recibido")
        print("=====================================")
        print("Proyecto:", data.get("cf_nombre_proyecto") or data.get("opportunity_name"))
        print("Contacto ID:", data.get("contact_id"))
        print("Oportunidad ID:", data.get("id"))
        print("=====================================\n")

        datos_excel = construir_datos_ficha_desde_webhook(data)

        nombre_limpio = limpiar_nombre_archivo(
            datos_excel.get("nombre_proyecto", "proyecto")
        )

        url_foto_edificio = obtener_primera_url(datos_excel.get("foto_edificio"))
        url_foto_montantes = obtener_primera_url(datos_excel.get("foto_montantes"))

        print("📸 URL FOTO EDIFICIO:", url_foto_edificio)
        print("📸 URL FOTO MONTANTES:", url_foto_montantes)

        foto_edificio_path = descargar_imagen(
            url_foto_edificio,
            f"{nombre_limpio}_foto_edificio"
        )

        foto_montantes_path = descargar_imagen(
            url_foto_montantes,
            f"{nombre_limpio}_foto_montantes"
        )

        datos_excel["foto_edificio_path"] = foto_edificio_path
        datos_excel["foto_montantes_path"] = foto_montantes_path

        archivo_generado = generar_excel_ficha_datos(datos_excel)

        print(f"✅ [FICHA DATOS WIN] Excel generado: {archivo_generado}")

        enviar_correo_ficha_datos_win(archivo_generado, datos_excel)
        # Después de enviar correo a WIN, mover automáticamente a Pendiente Inicio de Habilitación
        opp_id = data.get("id") or datos_excel.get("_opp_id")
        mover_oportunidad_a_pendiente_inicio_habilitacion(opp_id)

        return jsonify({
            "status": "ok",
            "mensaje": "Excel generado y correo enviado correctamente",
            "archivo": archivo_generado
        }), 200

    except Exception as e:
        print("❌ [FICHA DATOS WIN] Error:", str(e))
        return jsonify({
            "status": "error",
            "mensaje": str(e)
        }), 500

# =========================================================
# RUTA 1: EL GÉNESIS (Formulario de Asignación)
# =========================================================
@app.route('/webhook-formulario1', methods=['POST'])
def webhook_formulario1():
    datos = request.json
    if not datos:
        return jsonify({"error": "No data"}), 400

    nombre_proyecto_raw = obtener_campo(datos, "cf_nombre_proyecto")
    if not nombre_proyecto_raw:
        return jsonify({"error": "Falta el nombre del proyecto"}), 400

    nombre_proyecto = str(nombre_proyecto_raw).strip().upper()
    custom_fields = extraer_custom_fields_para_ghl(datos)

    ejecutivo_str = str(obtener_campo(datos, "cf_ejecutivo_principal")).upper()
    tipo_ingreso = str(obtener_campo(datos, "cf_tipo_ingreso")).capitalize()
    distrito = str(obtener_campo(datos, "cf_distrito")).capitalize()

    print(f"🚀 [FORM 1] Iniciando creación Génesis para: {nombre_proyecto}")

    # --- ESCUDO ANTI-DUPLICADOS ---
    res_search = requests.get(
        f"https://services.leadconnectorhq.com/opportunities/search?location_id={LOCATION_ID}&q={nombre_proyecto}",
        headers=HEADERS_GHL
    )
    if res_search.status_code == 200:
        resultados = res_search.json().get("opportunities", [])
        opp_existente = next((opp for opp in resultados if opp.get("pipelineId") == PIPELINE_ID), None)
        if opp_existente:
            print(f"⚠️ [FORM 1] El proyecto '{nombre_proyecto}' ya existe. Evitando creación duplicada.")
            return jsonify({"status": "ignored", "message": "El proyecto ya existe"}), 200

    owner_id = USERS_GHL["JEAN"] if "JEAN" in ejecutivo_str else USERS_GHL["YASMIN"]

    tags_contacto = [
        "NUEVO HUNTING",
        f"Origen: {tipo_ingreso}" if tipo_ingreso != "None" else "Origen: No Especificado",
        f"Distrito: {distrito}" if distrito != "None" else "Distrito: No Especificado",
        "Hunter: Jean" if owner_id == USERS_GHL["JEAN"] else "Hunter: Yasmin",
        "Trigger_Notificacion_BO"
    ]

    # 1. ATRAPAR Y TRANSFORMAR AL CONTACTO FANTASMA NATIVO
    contact_id = datos.get("contact_id") or datos.get("contact", {}).get("id")

    if contact_id:
        res_contact = requests.put(
            f"https://services.leadconnectorhq.com/contacts/{contact_id}",
            headers=HEADERS_GHL,
            json={
                "firstName": "Administración:",
                "lastName": nombre_proyecto,
                "assignedTo": owner_id,
                "tags": tags_contacto,
                "customFields": custom_fields
            }
        )
    else:
        res_contact = requests.post(
            "https://services.leadconnectorhq.com/contacts/",
            headers=HEADERS_GHL,
            json={
                "locationId": LOCATION_ID,
                "firstName": "Administración:",
                "lastName": nombre_proyecto,
                "assignedTo": owner_id,
                "tags": tags_contacto,
                "customFields": custom_fields
            }
        )
        contact_id = res_contact.json().get("contact", {}).get("id")

    if res_contact.status_code not in [200, 201]:
        print(f"❌ Error al procesar contacto: {res_contact.text}")
        return jsonify({"error": "Fallo al procesar contacto"}), 500

    # 2. Crear Company (El activo físico)
    res_company = requests.post(
        "https://services.leadconnectorhq.com/companies/",
        headers=HEADERS_GHL,
        json={
            "locationId": LOCATION_ID,
            "name": nombre_proyecto
        }
    )
    company_id = res_company.json().get("company", {}).get("id") if res_company.status_code in [200, 201] else None

    # 3. Vincular Company con Contacto
    if company_id:
        requests.post(
            f"https://services.leadconnectorhq.com/contacts/{contact_id}/companies/{company_id}",
            headers=HEADERS_GHL
        )

    # 4. Crear Oportunidad (SIN el array de followers aquí)
    res_opp = requests.post(
        "https://services.leadconnectorhq.com/opportunities/",
        headers=HEADERS_GHL,
        json={
            "locationId": LOCATION_ID,
            "contactId": contact_id,
            "pipelineId": PIPELINE_ID,
            "pipelineStageId": STAGE_FORM_1,
            "name": nombre_proyecto,
            "status": "open",
            "assignedTo": owner_id,
            "customFields": custom_fields
        }
    )

    if res_opp.status_code not in [200, 201]:
        print(f"❌ Error al crear la oportunidad en GHL: {res_opp.text}")
        return jsonify({"error": "Fallo al crear oportunidad"}), 500

    # 5. Agregar el Seguidor (Follower) en la ruta correcta
    opp_id = res_opp.json().get("opportunity", {}).get("id")
    if opp_id:
        res_follower = requests.post(
            f"https://services.leadconnectorhq.com/opportunities/{opp_id}/followers",
            headers=HEADERS_GHL,
            json={
                "followers": [USERS_GHL["STEFANO_BO"]]
            }
        )
        if res_follower.status_code not in [200, 201]:
            print(f"⚠️ Aviso: No se pudo agregar el seguidor. {res_follower.text}")

    print(f"✅ [FORM 1] Setup completado. Asignado a: {'Jean' if owner_id == USERS_GHL['JEAN'] else 'Yasmin'}")
    return jsonify({"status": "success"}), 200


# =========================================================
# RUTA 2: LA ACTUALIZACIÓN (Formulario de Ficha de Datos)
# =========================================================
@app.route('/webhook-formulario2', methods=['POST'])
def webhook_formulario2():
    datos = request.json
    if not datos: return jsonify({"error": "No data"}), 400

    nombre_proyecto_raw = obtener_campo(datos, "cf_nombre_proyecto")
    if not nombre_proyecto_raw:
        return jsonify({"error": "Falta el nombre del proyecto"}), 400

    nombre_proyecto = str(nombre_proyecto_raw).strip().upper()
    print(f"🔍 [FORM 2] Buscando oportunidad para actualizar: '{nombre_proyecto}'")

    # Búsqueda de Oportunidad
    res_search = requests.get(
        f"https://services.leadconnectorhq.com/opportunities/search?location_id={LOCATION_ID}&q={nombre_proyecto}",
        headers=HEADERS_GHL
    )
    resultados = res_search.json().get("opportunities", []) if res_search.status_code == 200 else []
    oportunidad_correcta = next((opp for opp in resultados if opp.get("pipelineId") == PIPELINE_ID), None)

    if not oportunidad_correcta:
        print(f"⚠️ [FORM 2] Oportunidad no encontrada.")
        return jsonify({"error": "Oportunidad no encontrada"}), 404

    if oportunidad_correcta.get("pipelineStageId") == STAGE_FORM_2:
        print(f"⚠️ [FORM 2] La ficha de '{nombre_proyecto}' ya fue procesada previamente.")
        return jsonify({"status": "ignored", "message": "Ficha ya procesada"}), 200

    opp_id = oportunidad_correcta.get("id")
    contacto_original_id = oportunidad_correcta.get("contactId")
    custom_fields_nuevos = extraer_custom_fields_para_ghl(datos)

    # 1. CONTACTO FANTASMA
    # OJO: no lo eliminamos todavía porque las fotos del formulario pueden depender de este contacto.
    contacto_fantasma_id = datos.get("contact_id") or datos.get("contact", {}).get("id")

    if contacto_fantasma_id and contacto_fantasma_id != contacto_original_id:
      print(f"👻 [FORM 2] Contacto fantasma detectado pero NO eliminado: {contacto_fantasma_id}")

    # 2. EXTRACCIÓN Y MAPEO ESPECÍFICO (Responsable y Dirección)
    resp_nombre = obtener_campo(datos, "cf_nombre_responsable_edificio") or ""
    resp_tel = obtener_campo(datos, "cf_telefono_responsable_edificio") or ""
    resp_correo = obtener_campo(datos, "cf_correo_responsable_edificio") or ""

    distrito = obtener_campo(datos, "cf_distrito") or ""
    urb = obtener_campo(datos, "cf_urbanizacion_edificio") or ""
    tipo_via = obtener_campo(datos, "cf_tipo_via") or ""
    nombre_via = obtener_campo(datos, "cf_nombre_via") or ""
    coords = obtener_campo(datos, "cf_coordenadas") or ""

    direccion_completa = f"{tipo_via} {nombre_via}, {urb}, {distrito} {coords}".strip(" ,")
    ciudad = obtener_campo(datos, "cf_departamento_edificio") or "Lima"
    estado = obtener_campo(datos, "cf_provincia_edificio") or "Lima"
    codigo_postal = obtener_campo(datos, "cf_codigo_postal_edificio") or ""

    # 3. ACTUALIZAR COMPANY
    # Obtener el companyId del contacto original para actualizarlo
    res_get_contact = requests.get(f"https://services.leadconnectorhq.com/contacts/{contacto_original_id}",
                                   headers=HEADERS_GHL)
    company_id = res_get_contact.json().get("contact", {}).get(
        "companyId") if res_get_contact.status_code == 200 else None

    if company_id:
        requests.put(
            f"https://services.leadconnectorhq.com/companies/{company_id}",
            headers=HEADERS_GHL,
            json={
                "locationId": LOCATION_ID,
                "name": nombre_proyecto,
                "phone": resp_tel,
                "address": direccion_completa,
                "city": ciudad,
                "state": estado,
                "postalCode": codigo_postal,
                "country": "Peru"
            }
        )

    # 4. ACTUALIZAR OPORTUNIDAD (Mover de etapa e inyectar CFs restantes)
    requests.put(
        f"https://services.leadconnectorhq.com/opportunities/{opp_id}",
        headers=HEADERS_GHL,
        json={
            "pipelineId": PIPELINE_ID,
            "pipelineStageId": STAGE_FORM_2,
            "name": nombre_proyecto,
            "customFields": custom_fields_nuevos
        }
    )

    # 5. ACTUALIZAR CONTACTO (Nativo + Custom Fields)
    requests.put(
        f"https://services.leadconnectorhq.com/contacts/{contacto_original_id}",
        headers=HEADERS_GHL,
        json={
            "firstName": resp_nombre if resp_nombre else "Administración:",
            "lastName": nombre_proyecto,
            "email": resp_correo,
            "phone": resp_tel,
            "customFields": custom_fields_nuevos
        }
    )
    # Guardar datos completos de la ficha para usarlos después en el Excel final
    guardar_datos_ficha_en_cache(
        nombre_proyecto=nombre_proyecto,
        opp_id=opp_id,
        contact_id=contacto_original_id,
        datos=datos
    )

    print(f"✅ [FORM 2] Proyecto '{nombre_proyecto}' actualizado. Empresa y Contacto sincronizados.")
    return jsonify({"status": "success"}), 200


# =========================================================
# FUNCIÓN INTERNA: ENVÍO DE CORREO SMTP
# =========================================================
def enviar_correo_win(datos_contacto):
    smtp_server = os.getenv("SMTP_SERVER")
    smtp_port = int(os.getenv("SMTP_PORT", 465))
    email_emisor = os.getenv("EMAIL_EMISOR")
    email_password = os.getenv("EMAIL_PASSWORD")
    email_destino = os.getenv("EMAIL_DESTINO")

    # CC fijo para correo de asignación
    email_cc_asignacion = os.getenv("EMAIL_CC_ASIGNACION")

    # Extracción segura de los campos
    tipo_predio = obtener_campo(datos_contacto, "cf_tipo_edificio") or "EDIFICIO"

    nombre_predio = (
        obtener_campo(datos_contacto, "cf_nombre_proyecto")
        or datos_contacto.get("opportunity_name", "NO ESPECIFICADO")
    )

    tipo_via = obtener_campo(datos_contacto, "cf_tipo_via") or ""
    nombre_via = obtener_campo(datos_contacto, "cf_nombre_via") or ""
    direccion = f"{tipo_via} {nombre_via}".strip() or "NO ESPECIFICADO"

    numeracion = obtener_campo(datos_contacto, "cf_numeracion_via") or "No especificado"
    distrito = obtener_campo(datos_contacto, "cf_distrito") or "NO ESPECIFICADO"
    coordenadas = obtener_campo(datos_contacto, "cf_coordenadas") or "No especificado"
    estreno = obtener_campo(datos_contacto, "cf_es_estreno") or "SI"
    inmobiliaria = obtener_campo(datos_contacto, "cf_inmobiliaria") or "NO ESPECIFICADO"

    ejecutivo = (
        obtener_campo(datos_contacto, "cf_ejecutivo_principal")
        or datos_contacto.get("owner", "NO ESPECIFICADO")
    )

    asignar_reasignar = obtener_campo(datos_contacto, "cf_tipo_gestion") or "ASIGNAR"

    fecha_actual = datetime.now().strftime("%d/%m/%Y")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"Solicitud de asignación - {str(nombre_predio).upper()}"
    msg["From"] = f"Vertical Futura <{email_emisor}>"
    msg["To"] = email_destino

    if email_cc_asignacion:
        msg["Cc"] = email_cc_asignacion

    html = f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.4;">
        <p>Estimados,</p>
        <p>Solicitamos la asignación del predio en mención.</p>
        <br>
        <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; max-width: 600px; border-color: #cccccc;">
          <tr style="background-color: #f2f2f2;">
            <th align="left" style="width: 40%; font-size: 14px;">FORMATO DE ASIGNACION</th>
            <th style="width: 60%;"></th>
          </tr>
          <tr><td><strong>Tipo de Predio:</strong></td><td>{str(tipo_predio).upper()}</td></tr>
          <tr><td><strong>Nombre del predio:</strong></td><td>{str(nombre_predio).upper()}</td></tr>
          <tr><td><strong>Direccion:</strong></td><td>{str(direccion).upper()}</td></tr>
          <tr><td><strong>Numeracion:</strong></td><td>{numeracion}</td></tr>
          <tr><td><strong>Distrito:</strong></td><td>{str(distrito).upper()}</td></tr>
          <tr><td><strong>Coordenadas:</strong></td><td>{coordenadas}</td></tr>
          <tr><td><strong>Cobertura Si/no:</strong></td><td>SI</td></tr>
          <tr><td><strong>Asignar/Reasignar:</strong></td><td>{str(asignar_reasignar).upper()}</td></tr>
          <tr><td><strong>Estreno:</strong></td><td>{str(estreno).upper()}</td></tr>
          <tr><td><strong>Fecha:</strong></td><td>{fecha_actual}</td></tr>
          <tr><td><strong>Inmobiliaria:</strong></td><td>{str(inmobiliaria).upper()}</td></tr>
          <tr><td><strong>Ejecutivo:</strong></td><td>{str(ejecutivo).upper()}</td></tr>
        </table>
        <p style="margin-top: 25px;">Saludos,</p>
        <p><strong>Stefano Sotomarino Goche</strong><br>Back Office - Futura</p>
      </body>
    </html>
    """

    msg.attach(MIMEText(html, "html"))

    try:
        destinatarios = [email_destino]

        if email_cc_asignacion:
            destinatarios.append(email_cc_asignacion)

        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(email_emisor, email_password)
            server.sendmail(email_emisor, destinatarios, msg.as_string())

        print(f"📧 Correo enviado a WIN para: {str(nombre_predio).upper()}")

        if email_cc_asignacion:
            print(f"📧 CC asignación enviado a: {email_cc_asignacion}")

        return True

    except Exception as e:
        print(f"❌ Error al enviar el correo a WIN: {e}")
        return False


# =========================================================
# RUTA 3: EL AUTO-AVANCE (Envío de Correo y Cambio de Etapa)
# =========================================================
@app.route('/webhook--enviar-correo-asignacion', methods=['POST'])
def webhook_enviar_correo():
    datos = request.json
    if not datos: return jsonify({"error": "No data"}), 400

    opp_id = datos.get("id")
    contact_id = datos.get("contact_id")

    # 1. RECUPERAR LOS CAMPOS PERSONALIZADOS QUE GHL OCULTA EN EL WEBHOOK LIGERO
    if contact_id:
        res_contact = requests.get(
            f"https://services.leadconnectorhq.com/contacts/{contact_id}",
            headers=HEADERS_GHL
        )
        if res_contact.status_code == 200:
            cfs = res_contact.json().get("contact", {}).get("customFields", [])
            for cf in cfs:
                # Inyectamos los campos ocultos directamente en nuestro diccionario de datos
                datos[cf["id"]] = cf.get("value")

    nombre_proyecto = obtener_campo(datos, "cf_nombre_proyecto") or datos.get("opportunity_name", "Desconocido")
    print(f"🚀 [FORM 3] Iniciando envío de correo WIN para: {nombre_proyecto}")

    # 2. Enviar el correo SMTP con todos los datos completos
    correo_enviado = enviar_correo_win(datos)

    # 3. Si el correo sale con éxito, mover la tarjeta automáticamente
    if correo_enviado and opp_id:
        STAGE_ESPERANDO = os.getenv("STAGE_ESPERANDO_WIN")

        res_update = requests.put(
            f"https://services.leadconnectorhq.com/opportunities/{opp_id}",
            headers=HEADERS_GHL,
            json={
                "pipelineId": os.getenv("PIPELINE_ID"),
                "pipelineStageId": STAGE_ESPERANDO
            }
        )
        if res_update.status_code == 200:
            print(f"✅ [FORM 3] Tarjeta movida automáticamente a 'Esperando Respuesta WIN'.")
            return jsonify({"status": "success", "message": "Correo enviado y tarjeta movida"}), 200
        else:
            print(f"❌ Error al mover la tarjeta: {res_update.text}")
            return jsonify({"status": "partial", "message": "Correo enviado pero fallo al mover la tarjeta"}), 500

    elif not correo_enviado:
        return jsonify({"status": "error", "message": "Fallo el servidor de correos"}), 500

    return jsonify({"status": "error", "message": "Faltan IDs para mover la tarjeta"}), 400


if __name__ == '__main__':
    app.run(port=5001, debug=True)