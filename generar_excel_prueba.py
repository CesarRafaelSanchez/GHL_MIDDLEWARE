from openpyxl import load_workbook
import os
import openpyxl.utils
import math

RUTA_PLANTILLA = "archivos_datos/plantillas/Ficha de Datos NOMBRE DEL PROYECTO.xlsx"
CARPETA_SALIDA = "salidas"

os.makedirs(CARPETA_SALIDA, exist_ok=True)


def escribir(ws, celda, valor):
    if valor is None:
        return
    valor = str(valor).strip()
    if valor == "":
        return
    ws[celda] = valor.upper()


def marcar(ws, celda, condicion):
    if condicion:
        ws[celda] = "X"


def parse_int_seguro(valor, default=0):
    if valor is None:
        return default
    try:
        return int(float(str(valor).strip()))
    except:
        return default


def parse_hogares_por_piso(valor_raw, num_pisos):
    if not valor_raw:
        return []
    
    partes = [p.strip() for p in str(valor_raw).split(",") if p.strip()]
    if len(partes) == 1:
        try:
            val = int(partes[0])
            return [val] * num_pisos
        except ValueError:
            return [partes[0]] * num_pisos
    else:
        resultado = []
        for p in partes:
            try:
                resultado.append(int(p))
            except ValueError:
                resultado.append(p)
        return resultado


def escribir_torres_dinamico(ws, datos):
    bloques = [
        {"name_cell": "A58", "header_row": 58, "homes_row": 59},
        {"name_cell": "A61", "header_row": 61, "homes_row": 62},
        {"name_cell": "A64", "header_row": 64, "homes_row": 65},
        {"name_cell": "A67", "header_row": 67, "homes_row": 68},
    ]

    torres_info = []
    for i in range(1, 4):
        nombre = datos.get(f"nombre_torre{i}")
        pisos = parse_int_seguro(datos.get(f"pisos_torre{i}"))
        hogares_raw = datos.get(f"hogares_por_piso_torre{i}")

        if not pisos and not hogares_raw:
            continue

        if not nombre:
            nombre = str(i)

        hogares_lista = parse_hogares_por_piso(hogares_raw, pisos)
        torres_info.append({
            "nombre": str(nombre).strip().upper(),
            "pisos": pisos,
            "hogares": hogares_lista
        })

    block_idx = 0
    for t in torres_info:
        nombre_torre = t["nombre"]
        num_floors = t["pisos"]
        homes_list = t["hogares"]

        needed_blocks = max(1, math.ceil(num_floors / 10))

        for b in range(needed_blocks):
            if block_idx >= len(bloques):
                print(f"⚠️ Se excedió el límite de 4 bloques. Omitiendo pisos de Torre {nombre_torre}.")
                break

            block = bloques[block_idx]
            ws[block["name_cell"]] = f"TORRE {nombre_torre}"

            start_floor = b * 10 + 1
            end_floor = (b + 1) * 10

            for col_idx in range(10):
                col_letter = openpyxl.utils.get_column_letter(3 + col_idx)
                floor_num = start_floor + col_idx

                header_cell = f"{col_letter}{block['header_row']}"
                ws[header_cell] = floor_num

                homes_cell = f"{col_letter}{block['homes_row']}"
                if floor_num <= num_floors:
                    idx = floor_num - 1
                    if idx < len(homes_list):
                        val = homes_list[idx]
                        try:
                            ws[homes_cell] = int(val)
                        except ValueError:
                            ws[homes_cell] = str(val).strip().upper()
                    else:
                        ws[homes_cell] = "N/A"
                else:
                    ws[homes_cell] = "N/A"

            block_idx += 1

    while block_idx < len(bloques):
        block = bloques[block_idx]
        ws[block["name_cell"]] = None
        for col_idx in range(10):
            col_letter = openpyxl.utils.get_column_letter(3 + col_idx)
            ws[f"{col_letter}{block['header_row']}"] = None
            ws[f"{col_letter}{block['homes_row']}"] = None
        block_idx += 1


def generar_excel_ficha_datos(datos):
    wb = load_workbook(RUTA_PLANTILLA)
    ws = wb["Ficha"]

    # DATOS GENERALES
    escribir(ws, "C5", datos.get("nombre_proyecto"))

    tipo_proyecto = (datos.get("tipo_proyecto") or "").upper()
    marcar(ws, "F8", tipo_proyecto == "NUEVO PREDIO")
    marcar(ws, "L8", tipo_proyecto == "AMPLIACION DE TORRE")

    fuente = (datos.get("fuente_origen") or "").upper()
    marcar(ws, "E12", fuente == "PROPIO")

    clasificacion = (datos.get("clasificacion") or "").upper()
    marcar(ws, "E16", clasificacion == "EDIFICIO")
    marcar(ws, "I16", clasificacion == "CONDOMINIO")

    tipo_construccion = (datos.get("tipo_construccion") or "").upper()
    marcar(ws, "E22", tipo_construccion == "ESTRENO")
    marcar(ws, "I22", tipo_construccion == "MODERNO")
    marcar(ws, "L22", tipo_construccion == "ANTIGUO")

    escribir(ws, "E23", datos.get("fecha_entrega_edificio"))
    escribir(ws, "E25", datos.get("fecha_termino_montantes"))
    escribir(ws, "E26", datos.get("fecha_termino_mecha"))

    junta = (datos.get("junta_directiva") or "").upper()
    marcar(ws, "E31", junta == "SI")
    marcar(ws, "I31", junta == "NO")

    escribir(ws, "D34", datos.get("cargo_responsable"))
    escribir(ws, "I34", datos.get("nombre_responsable"))
    escribir(ws, "D35", datos.get("telefono_responsable"))
    escribir(ws, "I35", datos.get("correo_responsable"))

    escribir(ws, "D39", datos.get("visita_inspeccion_tecnica"))
    horario = (datos.get("rango_horario_visita") or "").upper().replace(" ", "")
    marcar(ws, "I39", "9AM" in horario or "9A12" in horario)
    marcar(ws, "L39", "1PM" in horario or "1A4" in horario)

    # Dirección
    escribir(ws, "C43", datos.get("departamento"))
    escribir(ws, "I43", datos.get("provincia"))
    escribir(ws, "C44", datos.get("distrito"))
    escribir(ws, "I44", datos.get("urbanizacion"))
    escribir(ws, "C45", datos.get("codigo_postal"))
    escribir(ws, "C46", datos.get("tipo_via"))
    escribir(ws, "C47", datos.get("nombre_via"))
    escribir(ws, "C48", datos.get("numero_via"))
    escribir(ws, "C49", datos.get("coordenadas"))

    # Datos técnicos
    escribir(ws, "C53", datos.get("total_torres"))
    escribir(ws, "C54", datos.get("total_hogares"))

    # Torres dinámico
    escribir_torres_dinamico(ws, datos)

    # Canal & Gestor
    escribir(ws, "C72", datos.get("nombre_canal"))
    escribir(ws, "C75", datos.get("gestor"))
    escribir(ws, "I75", datos.get("celular_gestor"))

    nombre_archivo = f"Ficha de Datos - {datos.get('nombre_proyecto', 'PROYECTO')}.xlsx"
    ruta_salida = os.path.join(CARPETA_SALIDA, nombre_archivo)

    wb.save(ruta_salida)
    return ruta_salida


if __name__ == "__main__":
    datos_prueba = {
        "nombre_proyecto": "EDIFICIO PRUEBA ALLAMANDA MODIFIED",
        "tipo_proyecto": "Nuevo Predio",
        "fuente_origen": "Propio",
        "clasificacion": "Edificio",
        "tipo_construccion": "Moderno",
        "fecha_entrega_edificio": "31/07/2026",
        "fecha_termino_montantes": "30/05/2026",
        "fecha_termino_mecha": "30/05/2026",
        "junta_directiva": "No",
        "cargo_responsable": "Ingeniero de Proyectos",
        "nombre_responsable": "Juan Garcia",
        "telefono_responsable": "987654321",
        "correo_responsable": "responsable@proyecto.com",
        "visita_inspeccion_tecnica": "05/06/2026",
        "rango_horario_visita": "9 AM A 12 M",
        "departamento": "Lima",
        "provincia": "Lima",
        "distrito": "Surco",
        "urbanizacion": "Las Casuarinas",
        "codigo_postal": "15023",
        "tipo_via": "Calle",
        "nombre_via": "Allamanda",
        "numero_via": "115",
        "coordenadas": "-12.117178,-76.975572",
        "total_torres": "2",
        "total_hogares": "17",
        
        "nombre_torre1": "1",
        "pisos_torre1": "12",
        "hogares_por_piso_torre1": "3,3,3,3,3,3,3,3,3,3,2,2",
        
        "nombre_torre2": "2",
        "pisos_torre2": "5",
        "hogares_por_piso_torre2": "3",
        
        "nombre_canal": "FUTURA",
        "gestor": "JEAN PIERRE SIHUE SILVA",
        "celular_gestor": "918371086",
    }

    archivo = generar_excel_ficha_datos(datos_prueba)
    print(f"Excel generado correctamente en: {archivo}")