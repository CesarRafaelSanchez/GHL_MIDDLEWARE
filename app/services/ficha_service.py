import os
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
from io import BytesIO
from dotenv import load_dotenv
from app.utils.helpers import obtener_session_con_retries

requests = obtener_session_con_retries()

#load_dotenv()

# =========================================================
# RUTAS ABSOLUTAS (A prueba de balas)
# =========================================================
# Esto encuentra la ruta de tu proyecto GHL_System automáticamente
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

CARPETA_SALIDAS = os.path.join(BASE_DIR, "storage", "exports")
CARPETA_TEMP = os.path.join(BASE_DIR, "storage", "temp")
ARCHIVO_CACHE_FICHAS = os.path.join(BASE_DIR, "storage", "cache_fichas_datos.json")
RUTA_PLANTILLA_FICHA_DATOS = os.path.join(BASE_DIR, "resources", "templates", "Ficha de Datos NOMBRE DEL PROYECTO - Junio.xlsx")

# Creamos las carpetas si no existen
os.makedirs(CARPETA_SALIDAS, exist_ok=True)
os.makedirs(CARPETA_TEMP, exist_ok=True)
# =========================================================


def normalizar_clave_proyecto(nombre):
    return str(nombre or "").strip().upper()


def cargar_cache_fichas():
    if not os.path.exists(ARCHIVO_CACHE_FICHAS):
        return {}
    try:
        with open(ARCHIVO_CACHE_FICHAS, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}


def guardar_cache_fichas(cache):
    with open(ARCHIVO_CACHE_FICHAS, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def guardar_datos_ficha_en_cache(nombre_proyecto, opp_id, contact_id, datos):
    from app.database import guardar_oportunidad_db
    
    # Mapeamos los datos simulando el flujo del webhook
    datos_mapeados = construir_datos_ficha_desde_webhook(datos)
    
    # Incorporamos los campos de imágenes que se guardaron en la caché
    if "foto_edificio_path" in datos:
        datos_mapeados["foto_edificio_path"] = datos["foto_edificio_path"]
    if "foto_montantes_path" in datos:
        datos_mapeados["foto_montantes_path"] = datos["foto_montantes_path"]
        
    guardar_oportunidad_db(opp_id, contact_id, datos_mapeados, datos)
    print(f"💾 [DB SAVE] Ficha guardada en SQLite para proyecto: {nombre_proyecto}")


def recuperar_datos_ficha_de_cache(data):
    from app.database import obtener_oportunidad_db
    
    opp_id = data.get("id")
    contact_id = data.get("contact_id") or data.get("contact", {}).get("id")
    nombre = data.get("cf_nombre_proyecto") or data.get("opportunity_name")
    
    row = obtener_oportunidad_db(opp_id=opp_id, contact_id=contact_id, nombre_proyecto=nombre)
    
    if row:
        print(f"✅ [DB READ] Datos recuperados de SQLite para Opp: {opp_id or row.get('oportunidad_id')}")
        
        # Deserializar raw_json original si existe
        raw_json_str = row.get("raw_json")
        base_dict = {}
        if raw_json_str:
            try:
                base_dict = json.loads(raw_json_str)
            except Exception:
                pass
                
        # Mapa de columnas DB a campos GHL para sobrescribir con ediciones manuales
        DB_COL_TO_GHL_KEY = {
            "nombre_proyecto": "cf_nombre_proyecto",
            "tipo_proyecto": "cf_tipo_proyecto",
            "fuente_origen": "cf_fuente_hunting",
            "clasificacion": "cf_clasificacion_proyecto",
            "tipo_construccion": "cf_tipo_construccion_edificio",
            "fecha_entrega_edificio": "cf_fecha_entrega_edificio_estreno",
            "fecha_termino_montantes": "cf_fecha_termino_montantes_edificio_estreno",
            "fecha_termino_mecha": "cf_fecha_termino_mecha_edificio_estreno",
            "junta_directiva": "cf_junta_directiva",
            "cargo_responsable": "cf_cargo_responsable_edificio",
            "nombre_responsable": "cf_nombre_responsable_edificio",
            "telefono_responsable": "cf_telefono_responsable_edificio",
            "correo_responsable": "cf_correo_responsable_edificio",
            "hogares_por_piso_torre1": "cf_hogares_por_piso_torre1",
            "hogares_por_piso_torre2": "cf_hogares_por_piso_torre2",
            "hogares_por_piso_torre3": "cf_hogares_por_piso_torre3",
            "visita_inspeccion_tecnica": "cf_visita_inspeccion_tecnica_win",
            "rango_horario_visita": "cf_rango_horario_visita_tecnica",
            "departamento": "cf_departamento_edificio",
            "provincia": "cf_provincia_edificio",
            "distrito": "cf_distrito",
            "urbanizacion": "cf_urbanizacion_edificio",
            "codigo_postal": "cf_codigo_postal_edificio",
            "tipo_via": "cf_tipo_via",
            "nombre_via": "cf_nombre_via",
            "numero_via": "cf_numeracion_via",
            "coordenadas": "cf_coordenadas",
            "total_torres": "cf_total_torres_proyecto",
            "total_hogares": "cf_total_hogares_proyecto",
            "nombre_torre1": "cf_nombre_torre1_proyecto",
            "pisos_torre1": "cf_pisos_torre1_proyecto",
            "hogares_torre1": "cf_hogares_torre1_proyecto",
            "nombre_torre2": "cf_nombre_torre2_proyecto",
            "pisos_torre2": "cf_pisos_torre2_proyecto",
            "hogares_torre2": "cf_hogares_torre2_proyecto",
            "nombre_torre3": "cf_nombre_torre3_proyecto",
            "pisos_torre3": "cf_pisos_torre3_proyecto",
            "hogares_torre3": "cf_hogares_torre3_proyecto",
            "clientes_interesados": "cf_cantidad_clientes_interesados",
            "nombre_canal": "cf_nombre_canal_hunting",
            "gestor": "cf_gestor_real",
            "foto_edificio": "cf_foto_edificio",
            "foto_montantes": "cf_foto_montantes"
        }
        
        # Sincronizamos las columnas de la DB al diccionario base
        for col_name, val in row.items():
            if col_name == "raw_json":
                continue
            # Mantener valor directo
            base_dict[col_name] = val
            
            # Mapear a campo GHL si existe en el mapa
            if col_name in DB_COL_TO_GHL_KEY:
                base_dict[DB_COL_TO_GHL_KEY[col_name]] = val
                
            if col_name == "celular_gestor" and val:
                if "user" not in base_dict or not isinstance(base_dict["user"], dict):
                    base_dict["user"] = {}
                base_dict["user"]["phone"] = val
                
        # Las fotos locales se copian directamente
        base_dict["foto_edificio_path"] = row.get("foto_edificio_path")
        base_dict["foto_montantes_path"] = row.get("foto_montantes_path")
        
        # Mezclamos con el payload de entrada (los datos de nuestra base de datos local sobreescriben al webhook de GHL)
        data_mezclada = dict(data)
        data_mezclada.update(base_dict)
        data_mezclada["id"] = opp_id or row.get("oportunidad_id")
        data_mezclada["contact_id"] = contact_id or row.get("contacto_id")
        
        return data_mezclada
        
    print("⚠️ [DB READ] No se encontraron datos para la oportunidad en SQLite.")
    return data


def escribir_excel(ws, celda, valor):
    if valor is None:
        return
    valor = str(valor).strip()
    if valor == "":
        return
    ws[celda] = valor.upper()


def marcar_excel(ws, celda, condicion):
    if condicion:
        ws[celda] = "X"


def obtener_primera_url(valor):
    if not valor:
        return ""

    if isinstance(valor, list) and len(valor) > 0:
        val = valor[0]
    else:
        val = valor

    if isinstance(val, dict):
        # Caso de diccionario anidado con UUID como clave: {'uuid': {'url': '...'}}
        if len(val) == 1 and isinstance(list(val.values())[0], dict):
            inner = list(val.values())[0]
            url = inner.get("url") or inner.get("value") or inner.get("downloadUrl")
            if url:
                return url
        url = val.get("url") or val.get("value") or val.get("downloadUrl")
        if url:
            return url

    if isinstance(val, str):
        return val.strip()

    return str(val)


def limpiar_nombre_archivo(nombre):
    nombre = str(nombre or "archivo")
    caracteres_invalidos = '<>:"/\\|?*'
    for c in caracteres_invalidos:
        nombre = nombre.replace(c, "_")
    return nombre.strip().replace(" ", "_")


# === TU FUNCIÓN ORIGINAL EXACTA DE DESCARGA ===
def descargar_imagen(url, nombre_archivo):
    if not url:
        print("⚠️ No llegó URL de imagen.")
        return None
    try:
        nombre_archivo = limpiar_nombre_archivo(nombre_archivo)
        token = os.getenv('GHL_ACCESS_TOKEN')
        location_id = os.getenv('GHL_LOCATION_ID')

        # EL PASE VIP: Agregamos alt=media y el locationId al final del link de GHL
        url = url.strip()
        separador = "&" if "?" in url else "?"
        url_final = f"{url}{separador}alt=media&locationId={location_id}"

        headers = {
            "Authorization": f"Bearer {token}",
            "Version": "2021-04-15",
            "Accept": "*/*"
        }

        print(f"📸 Intentando descargar imagen GHL con Pase VIP: {url_final}")
        response = requests.get(url_final, headers=headers, timeout=30, allow_redirects=True)

        print("📸 Status descarga imagen:", response.status_code)

        if response.status_code != 200:
            print(f"❌ Error descargando imagen ({response.status_code}): {response.text[:200]}")
            return None

        # Convertimos la imagen asegurando compatibilidad con Excel
        imagen = PILImage.open(BytesIO(response.content))
        if imagen.mode in ("RGBA", "P"):
            imagen = imagen.convert("RGB")
        else:
            imagen = imagen.convert("RGB")

        ruta = os.path.join(CARPETA_TEMP, f"{nombre_archivo}.jpg")
        imagen.save(ruta, format="JPEG", quality=90)

        print(f"✅ Imagen descargada y convertida correctamente: {ruta}")
        return ruta

    except Exception as e:
        print(f"❌ Error descargando/convirtiendo imagen GHL: {e}")
        return None

def insertar_imagen_excel(ws, ruta_imagen, celda, ancho=330, alto=260):
    if not ruta_imagen:
        print(f"[WARN] No hay imagen para insertar en {celda}")
        return
    if not os.path.exists(ruta_imagen):
        print(f"[WARN] No existe la imagen local: {ruta_imagen}")
        return
    try:
        img = ExcelImage(ruta_imagen)
        img.width = ancho
        img.height = alto
        ws.add_image(img, celda)
        print(f"[OK] Imagen insertada en Excel: {ruta_imagen} -> {celda}")
    except Exception as e:
        print(f"[ERROR] Error insertando imagen en Excel: {e}")


def parsear_hogares_por_piso(valor, num_pisos):
    if not num_pisos:
        return []
    try:
        num_pisos = int(num_pisos)
    except:
        return []

    if not valor:
        return [0] * num_pisos

    valor_str = str(valor).strip()
    # Si es un solo entero
    if valor_str.isdigit():
        return [int(valor_str)] * num_pisos

    # Si es una lista separada por comas
    try:
        parts = []
        for x in valor_str.split(','):
            x_clean = x.strip()
            if x_clean.isdigit():
                parts.append(int(x_clean))
            else:
                parts.append(0)
        if len(parts) < num_pisos:
            parts += [0] * (num_pisos - len(parts))
        elif len(parts) > num_pisos:
            parts = parts[:num_pisos]
        return parts
    except:
        return [0] * num_pisos


def generar_excel_ficha_datos(datos):
    wb = load_workbook(RUTA_PLANTILLA_FICHA_DATOS)
    ws = wb["Ficha"]

    escribir_excel(ws, "C5", datos.get("nombre_proyecto"))
    tipo_proyecto = (datos.get("tipo_proyecto") or "").upper()
    marcar_excel(ws, "F8", tipo_proyecto == "NUEVO PREDIO")
    marcar_excel(ws, "L8", tipo_proyecto == "AMPLIACION DE TORRE")
    fuente = (datos.get("fuente_origen") or "").upper()
    marcar_excel(ws, "E12", fuente == "PROPIO")
    clasificacion = (datos.get("clasificacion") or "").upper()
    marcar_excel(ws, "E16", clasificacion == "EDIFICIO")
    marcar_excel(ws, "I16", clasificacion == "CONDOMINIO")
    tipo_construccion = (datos.get("tipo_construccion") or "").upper()
    marcar_excel(ws, "E22", tipo_construccion == "ESTRENO")
    marcar_excel(ws, "I22", tipo_construccion == "MODERNO")
    marcar_excel(ws, "L22", tipo_construccion == "ANTIGUO")
    escribir_excel(ws, "E23", datos.get("fecha_entrega_edificio"))
    escribir_excel(ws, "E25", datos.get("fecha_termino_montantes"))
    escribir_excel(ws, "E26", datos.get("fecha_termino_mecha"))
    junta = (datos.get("junta_directiva") or "").upper()
    marcar_excel(ws, "E31", junta == "SI")
    marcar_excel(ws, "I31", junta == "NO")
    escribir_excel(ws, "D34", datos.get("cargo_responsable"))
    escribir_excel(ws, "I34", datos.get("nombre_responsable"))
    escribir_excel(ws, "D35", datos.get("telefono_responsable"))
    escribir_excel(ws, "I35", datos.get("correo_responsable"))

    escribir_excel(ws, "D39", datos.get("visita_inspeccion_tecnica"))
    horario = (datos.get("rango_horario_visita") or "").upper()
    marcar_excel(ws, "I39", horario in ["9 AM A 12 AM", "9AM A 12M", "9 AM A 12M"])
    marcar_excel(ws, "L39", horario == "1 PM A 4 PM")

    escribir_excel(ws, "C43", datos.get("departamento"))
    escribir_excel(ws, "I43", datos.get("provincia"))
    escribir_excel(ws, "C44", datos.get("distrito"))
    escribir_excel(ws, "I44", datos.get("urbanizacion"))
    escribir_excel(ws, "C45", datos.get("codigo_postal"))
    escribir_excel(ws, "C46", datos.get("tipo_via"))
    escribir_excel(ws, "C47", datos.get("nombre_via"))
    escribir_excel(ws, "C48", datos.get("numero_via"))
    escribir_excel(ws, "C49", datos.get("coordenadas"))

    # Lógica de torres y hogares por piso dinámica
    escribir_excel(ws, "C53", datos.get("total_torres"))
    escribir_excel(ws, "C54", datos.get("total_hogares"))

    towers = []
    for i in range(1, 4):
        t_name = datos.get(f"nombre_torre{i}")
        t_pisos = datos.get(f"pisos_torre{i}")
        t_hogares_piso = datos.get(f"hogares_por_piso_torre{i}")

        if t_pisos:
            try:
                npisos = int(t_pisos)
            except:
                npisos = 0
            if npisos > 0:
                t_name_val = t_name if t_name else str(i)
                hogares_list = parsear_hogares_por_piso(t_hogares_piso, npisos)
                towers.append({
                    "name": t_name_val,
                    "pisos": npisos,
                    "hogares_list": hogares_list
                })

    bloques_filas = [58, 61, 64, 67]
    bloque_idx = 0

    for t in towers:
        pisos_restantes = t["pisos"]
        hogares_restantes = list(t["hogares_list"])
        tower_block_num = 0

        while pisos_restantes > 0 and bloque_idx < len(bloques_filas):
            row_start = bloques_filas[bloque_idx]

            if tower_block_num == 0:
                escribir_excel(ws, f"A{row_start}", f"TORRE {t['name']}")
            else:
                escribir_excel(ws, f"A{row_start}", "")

            escribir_excel(ws, f"B{row_start}", "PISO :")
            escribir_excel(ws, f"B{row_start+1}", "HOGARES POR PISO")

            for c in range(1, 11):
                col_letter = get_column_letter(2 + c)
                floor_in_block = tower_block_num * 10 + c

                escribir_excel(ws, f"{col_letter}{row_start}", floor_in_block)

                if floor_in_block <= t["pisos"]:
                    h_val = hogares_restantes[floor_in_block - 1]
                    escribir_excel(ws, f"{col_letter}{row_start+1}", h_val)
                else:
                    escribir_excel(ws, f"{col_letter}{row_start+1}", "N/A")

            escribir_excel(ws, f"M{row_start}", "TOTAL")
            ws[f"M{row_start+1}"] = f"=SUM(C{row_start+1}:L{row_start+1})"

            pisos_restantes -= 10
            tower_block_num += 1
            bloque_idx += 1

    escribir_excel(ws, "C72", datos.get("nombre_canal"))
    escribir_excel(ws, "C75", datos.get("gestor"))
    escribir_excel(ws, "I75", datos.get("celular_gestor"))

    insertar_imagen_excel(ws, datos.get("foto_edificio_path"), "A80", 330, 260)
    insertar_imagen_excel(ws, datos.get("foto_montantes_path"), "F80", 330, 260)

    nombre_proyecto = limpiar_nombre_archivo(datos.get("nombre_proyecto", "PROYECTO")).replace("_", " ")
    nombre_archivo = f"Ficha de Datos - {nombre_proyecto}.xlsx"
    ruta_salida = os.path.join(CARPETA_SALIDAS, nombre_archivo)
    wb.save(ruta_salida)
    return ruta_salida


def formatear_fecha(fecha):
    if not fecha: return ""
    try:
        return datetime.strptime(str(fecha), "%Y-%m-%d").strftime("%d/%m/%Y")
    except:
        return str(fecha)


def construir_datos_ficha_desde_webhook(data):
    datos = {
        "nombre_proyecto": data.get("cf_nombre_proyecto") or data.get("opportunity_name"),
        "tipo_proyecto": data.get("cf_tipo_proyecto"),
        "fuente_origen": data.get("cf_fuente_hunting"),
        "clasificacion": "Condominio" if "condominio" in str(
            data.get("cf_clasificacion_proyecto", "")).lower() else "Edificio",
        "tipo_construccion": (
            "Estreno" if "estreno" in str(data.get("cf_tipo_construccion_edificio", "")).lower()
            else "Moderno" if "moderno" in str(data.get("cf_tipo_construccion_edificio", "")).lower()
            else "Antiguo" if "antiguo" in str(data.get("cf_tipo_construccion_edificio", "")).lower()
            else data.get("cf_tipo_construccion_edificio")
        ),
        "fecha_entrega_edificio": formatear_fecha(
            data.get("cf_fecha_entrega_edificio_estreno") or data.get("cf_fecha_entrega_edificio")),
        "fecha_termino_montantes": formatear_fecha(data.get("cf_fecha_termino_montantes_edificio_estreno")),
        "fecha_termino_mecha": formatear_fecha(data.get("cf_fecha_termino_mecha_edificio_estreno")),
        "junta_directiva": data.get("cf_junta_directiva"),
        "cargo_responsable": data.get("cf_cargo_responsable_edificio") or data.get("cf_cargo_responsable"),
        "nombre_responsable": data.get("cf_nombre_responsable_edificio"),
        "telefono_responsable": data.get("cf_telefono_responsable_edificio"),
        "correo_responsable": data.get("cf_correo_responsable_edificio"),
        "hogares_por_piso_torre1": data.get("cf_hogares_por_piso_torre1"),
        "hogares_por_piso_torre2": data.get("cf_hogares_por_piso_torre2"),
        "hogares_por_piso_torre3": data.get("cf_hogares_por_piso_torre3"),
        "visita_inspeccion_tecnica": formatear_fecha(data.get("cf_visita_inspeccion_tecnica_win")),
        "rango_horario_visita": data.get("cf_rango_horario_visita_tecnica"),
        "departamento": data.get("cf_departamento_edificio"),
        "provincia": data.get("cf_provincia_edificio"),
        "distrito": data.get("cf_distrito"),
        "urbanizacion": data.get("cf_urbanizacion_edificio"),
        "codigo_postal": data.get("cf_codigo_postal_edificio"),
        "tipo_via": data.get("cf_tipo_via"),
        "nombre_via": data.get("cf_nombre_via"),
        "numero_via": data.get("cf_numeracion_via"),
        "coordenadas": data.get("cf_coordenadas"),
        "total_torres": data.get("cf_total_torres_proyecto"),
        "total_hogares": data.get("cf_total_hogares_proyecto"),
        "nombre_torre1": data.get("cf_nombre_torre1_proyecto"),
        "pisos_torre1": data.get("cf_pisos_torre1_proyecto"),
        "hogares_torre1": data.get("cf_hogares_torre1_proyecto"),
        "nombre_torre2": data.get("cf_nombre_torre2_proyecto"),
        "pisos_torre2": data.get("cf_pisos_torre2_proyecto"),
        "hogares_torre2": data.get("cf_hogares_torre2_proyecto"),
        "nombre_torre3": data.get("cf_nombre_torre3_proyecto"),
        "pisos_torre3": data.get("cf_pisos_torre3_proyecto"),
        "hogares_torre3": data.get("cf_hogares_torre3_proyecto"),
        "clientes_interesados": data.get("cf_cantidad_clientes_interesados"),
        "nombre_canal": data.get("cf_nombre_cancal_hunting") or data.get("cf_nombre_canal_hunting"),
        "gestor": data.get("cf_gestor_real"),
        "celular_gestor": data.get("user", {}).get("phone", ""),
        "foto_edificio": obtener_primera_url(data.get("cf_foto_edificio")),
        "foto_montantes": obtener_primera_url(data.get("cf_foto_montantes")),
    }
    print("[MAPEO] cf_foto_edificio:", data.get("cf_foto_edificio"))
    print("[MAPEO] cf_foto_montantes:", data.get("cf_foto_montantes"))
    return datos